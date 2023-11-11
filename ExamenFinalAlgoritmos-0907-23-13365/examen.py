import openpyxl
import argparse
import tkinter as tk
from tkinter import messagebox, simpledialog

book = openpyxl.load_workbook('vehiculos.xlsx')
sheet = book.active

def listar_vehiculos(text_widget):

    text_widget.delete('1.0', tk.END)

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        text_widget.insert(tk.END, f"{header.ljust(12)}")
    text_widget.insert(tk.END, '\n')

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                text_widget.insert(tk.END, f"{str(data[i]).ljust(12)}")
            else:
                text_widget.insert(tk.END, "".ljust(12))
        text_widget.insert(tk.END, '\n')

def crear_vehiculo(text_widget):

    codigo = simpledialog.askstring("Input", "Ingrese el código del vehiculo:")
    marca = simpledialog.askstring("Input", "Ingrese la marca del vehiculo:")
    modelo = simpledialog.askstring("Input", "Ingrese el modelo del producto:")
    precio = simpledialog.askfloat("Input", "Ingrese el precio del vehiculo:")
    kilometraje = simpledialog.askinteger("Input", "Ingrese el kilometraje del vehiculo:")

    data = [codigo, marca, modelo, precio, kilometraje]
    sheet.append(data)

    book.save('vehiculos.xlsx')

def eliminar_vehiculo(text_widget):
    try:
        codigo = simpledialog.askstring("Input", "Ingrese el código del vehiculo que desea eliminar:")

        vehiculo_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el vehiculo\n")
                vehiculo_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino)
                text_widget.insert(tk.END, "El vehiculo ha sido eliminado.\n")
                break 

        if not vehiculo_encontrado:
            text_widget.insert(tk.END, "No se encontró el vehiculo con el código proporcionado\n")
            

        book.save('vehiculos.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

def editar_vehiculo(text_widget):

    codigo = simpledialog.askstring("Input", "Ingrese el codigo del vehiculo que desea actualizar")

    vehiculo_encontrado = False

    for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el producto\n")
                vehiculo_encontrado = True
                editar = simpledialog.askstring("Input", "¿Que desea editar? (Marca, Modelo, Precio, Kilometraje)")
                if editar == "Marca":
                    fila_destino = celda.row
                    columna_destino = 2
                    celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                    marca_nueva = simpledialog.askstring("Input", "Ingrese la nueva marca:")
                    celda_destino.value = marca_nueva
                    break
                elif editar == "Modelo":
                    fila_destino = celda.row
                    columna_destino = 3
                    celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                    modelo_nuevo = simpledialog.askstring("Input", "Ingrese el nuevo modelo:")
                    celda_destino.value = modelo_nuevo
                    break
                elif editar == "Precio":
                    fila_destino = celda.row
                    columna_destino = 4
                    celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                    precio_nuevo = simpledialog.askfloat("Input", "Ingrese el nuevo precio:")
                    celda_destino.value = precio_nuevo
                    break
                elif editar == "Kilometraje":
                    fila_destino = celda.row
                    columna_destino = 5
                    celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                    kilometraje_nuevo = simpledialog.askinteger("Input", "Ingrese el nuevo kilometraje:")
                    celda_destino.value = kilometraje_nuevo
                    break
                else:
                    text_widget.insert(tk.END, "Función no valida\n")

    if not vehiculo_encontrado:
        text_widget.insert(tk.END, "No se encontró el producto con el código proporcionado\n")

        book.save('vehiculos.xlsx')
    
def main():
    root = tk.Tk()
    root.title("Inventario")

    frame1 = tk.Frame(root)
    frame1.pack()

    frame2 = tk.Frame(root)
    frame2.pack()

    text_widget = tk.Text(frame2)
    text_widget.pack()

    menu = tk.Menu(root)
    root.config(menu=menu)

    vehiculos_menu = tk.Menu(menu)
    menu.add_cascade(label="Mantenimiento de vehiculos", menu=vehiculos_menu)

    vehiculos_menu.add_command(label='Listar Vehiculos', command=lambda: listar_vehiculos(text_widget))
    vehiculos_menu.add_command(label='Crear Vehiculos', command=lambda: crear_vehiculo(text_widget))
    vehiculos_menu.add_command(label='Editar Vehiculo', command=lambda: editar_vehiculo(text_widget))
    vehiculos_menu.add_command(label='Eliminar Vehiculo', command=lambda: eliminar_vehiculo(text_widget))
    

    root.mainloop()

if __name__ == "__main__":
    main()